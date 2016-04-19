#!/Python27/python
from openpyxl import load_workbook
import MySQLdb

# Open the workbook and define the worksheet
wb = load_workbook(filename = 'TabulatedSLA.xlsx')
sheet = wb['Sheet1']

# Establish a MySQL connection
database = MySQLdb.connect (host="localhost", user = "root", passwd = "toor", db = "cloud")

# Get the cursor, which is used to traverse the database, line by line
cursor = database.cursor()

empty_table = "delete from sla;delete from rank"

cursor.execute(empty_table)
cursor.close()

cursor = database.cursor()
# Create the INSERT INTO sql query
query = """INSERT INTO sla VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""

# Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
for r in range(1, len(sheet.rows)):
      product      = sheet.cell(row=r,column=1).value
      customer = sheet.cell(row=r,column=2).value
      rep          = sheet.cell(row=r,column=3).value
      date     = sheet.cell(row=r,column=4).value
      actual       = sheet.cell(row=r,column=5).value
      expected = sheet.cell(row=r,column=6).value
      open        = sheet.cell(row=r,column=7).value
      closed       = sheet.cell(row=r,column=8).value
      city     = sheet.cell(row=r,column=9).value
      state        = sheet.cell(row=r,column=10).value
      zip         = sheet.cell(row=r,column=11).value
      pop          = sheet.cell(row=r,column=12).value
      region   = sheet.cell(row=r,column=13).value
      # Assign values from each row
      values = (product, customer, rep, date, actual, expected, open, closed, city, state, zip, pop, region)

      # Execute sql Query
      cursor.execute(query, values)

# Close the cursor
wb = load_workbook(filename = 'TabulatedSLA.xlsx')
sheet = wb['Sheet2']


# Get the cursor, which is used to traverse the database, line by line
cursor = database.cursor()

# Create the INSERT INTO sql query
query = """INSERT INTO rank VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""

# Create a For loop to iterate through each row in the XLS file, starting at row 2 to skip the headers
for r in range(1, len(sheet.rows)):
      product      = sheet.cell(row=r,column=1).value
      customer = sheet.cell(row=r,column=2).value
      rep          = sheet.cell(row=r,column=3).value
      date     = sheet.cell(row=r,column=4).value
      actual       = sheet.cell(row=r,column=5).value
      expected = sheet.cell(row=r,column=6).value
      open        = sheet.cell(row=r,column=7).value
      closed       = sheet.cell(row=r,column=8).value
      city     = sheet.cell(row=r,column=9).value
      state        = sheet.cell(row=r,column=10).value
      zip         = sheet.cell(row=r,column=11).value
      pop          = sheet.cell(row=r,column=12).value
      region   = sheet.cell(row=r,column=13).value
      #print date,'\n'
      # Assign values from each row
      values = (product, customer, rep, date, actual, expected, open, closed, city, state, zip, pop, region)

      # Execute sql Query
      cursor.execute(query, values)

# Close the cursor
cursor.close()

# Commit the transaction
database.commit()

# Close the database connection
database.close()

